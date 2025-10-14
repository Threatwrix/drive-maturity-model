#!/usr/bin/env python3
"""
Generate all 52 1Secure YAML check files directly from the Excel source.
Maps to DRIVE maturity levels 1-5 based on severity thresholds.
Binary advancement model - no points, just pass/fail criteria.
"""

import yaml
import csv
from datetime import datetime
from openpyxl import load_workbook

def map_risk_to_levels(category, metric, measure_type, low, medium, high):
    """
    Map 1Secure risk thresholds to DRIVE maturity levels.

    Key principle: Higher severity = lower maturity level (more critical).
    - Level 1 (Critical Exposure): Immediate threats
    - Level 2 (High Risk): Short-term threats
    - Level 3 (Baseline): Standard controls
    - Level 4 (Enhanced): Proactive management
    - Level 5 (State-of-the-Art): Continuous excellence
    """

    levels = []

    # Binary risks - single level assignment
    if measure_type == 'Binary':
        if 'Third-Party' in metric or 'MS Graph' in metric:
            levels.append({
                'level': 2,
                'condition': f'{metric} detected',
                'severity': 'High',
                'timeline': 'weeks'
            })
        elif 'Password Not Required' in metric or 'Dangerous Default' in metric or 'Global Administrator' in metric or 'Kerberoast' in metric or 'SID History' in metric or 'RPC Coercion' in metric or 'DC Logon' in metric or 'ESC' in metric or 'Obsolete.*DC' in metric or 'SMB v1' in metric:
            levels.append({
                'level': 1,
                'condition': f'{metric} detected',
                'severity': 'Critical',
                'timeline': 'immediate to days'
            })
        elif 'Audit Log' in metric or 'Conditional Access' in metric or 'Domain Registration' in metric:
            levels.append({
                'level': 2,
                'condition': f'{metric} not configured',
                'severity': 'High',
                'timeline': 'days to weeks'
            })
        else:
            levels.append({
                'level': 3,
                'condition': f'{metric} detected',
                'severity': 'Medium',
                'timeline': 'weeks to months'
            })

    # Numeric/Percentage risks - multi-level thresholds
    else:
        # Sensitive data risks (most critical)
        if 'Sensitive' in metric:
            if high and high != '-':
                levels.append({
                    'level': 1,
                    'condition': f'≥{high} threshold for {metric}',
                    'severity': 'Critical',
                    'timeline': 'immediate to hours'
                })
            if medium and medium != '-':
                levels.append({
                    'level': 2,
                    'condition': f'{medium} threshold for {metric}',
                    'severity': 'High',
                    'timeline': 'days to weeks'
                })
            if low and low not in ['No risk', '-', 'Below 0']:
                levels.append({
                    'level': 3,
                    'condition': f'<{medium.split()[0] if medium != "-" else "threshold"} for {metric}',
                    'severity': 'Medium',
                    'timeline': 'weeks'
                })

        # Identity risks (critical to high)
        elif category == 'Identity':
            if high and high != '-' and 'Password' in metric:
                levels.append({
                    'level': 1,
                    'condition': f'≥{high} for {metric}',
                    'severity': 'Critical',
                    'timeline': 'days'
                })
            elif high and high != '-':
                levels.append({
                    'level': 2,
                    'condition': f'≥{high} for {metric}',
                    'severity': 'High',
                    'timeline': 'weeks'
                })

            if medium and medium != '-':
                target_level = 2 if 'Password' in metric else 3
                levels.append({
                    'level': target_level,
                    'condition': f'{medium} for {metric}',
                    'severity': 'High' if target_level == 2 else 'Medium',
                    'timeline': 'weeks' if target_level == 2 else 'months'
                })

        # Infrastructure risks (mostly baseline)
        elif category == 'Infrastructure':
            # ADCS ESC vulnerabilities are critical
            if 'ESC' in metric:
                levels.append({
                    'level': 1,
                    'condition': f'Any {metric} vulnerabilities',
                    'severity': 'Critical',
                    'timeline': 'immediate'
                })
            # Obsolete DCs are critical
            elif 'Obsolete' in metric and 'Domain Controller' in metric:
                levels.append({
                    'level': 1,
                    'condition': f'Any {metric}',
                    'severity': 'Critical',
                    'timeline': 'days'
                })
            # Other obsolete systems are high risk
            elif 'Obsolete' in metric or 'SMBv1' in metric:
                if high and high != '-':
                    levels.append({
                        'level': 2,
                        'condition': f'≥{high} for {metric}',
                        'severity': 'High',
                        'timeline': 'weeks'
                    })
                if medium and medium != '-':
                    levels.append({
                        'level': 3,
                        'condition': f'{medium} for {metric}',
                        'severity': 'Medium',
                        'timeline': 'months'
                    })
            # Disabled/inactive resources are baseline hygiene
            elif 'Disabled' in metric or 'Inactive' in metric or 'Empty' in metric:
                if high and high != '-':
                    levels.append({
                        'level': 3,
                        'condition': f'≥{high} for {metric}',
                        'severity': 'Medium',
                        'timeline': 'weeks'
                    })
                if medium and medium != '-':
                    levels.append({
                        'level': 4,
                        'condition': f'{medium} for {metric}',
                        'severity': 'Low',
                        'timeline': 'months'
                    })
            else:
                # Default infrastructure to Level 2-3
                if high and high != '-':
                    levels.append({
                        'level': 2,
                        'condition': f'≥{high} for {metric}',
                        'severity': 'High',
                        'timeline': 'weeks'
                    })
                if medium and medium != '-':
                    levels.append({
                        'level': 3,
                        'condition': f'{medium} for {metric}',
                        'severity': 'Medium',
                        'timeline': 'months'
                    })

        # Data risks (access control, sharing)
        else:
            # Broken inheritance is baseline hygiene
            if 'Broken' in metric or 'Inheritance' in metric:
                if high and high != '-':
                    levels.append({
                        'level': 3,
                        'condition': f'≥{high} for {metric}',
                        'severity': 'Medium',
                        'timeline': 'weeks'
                    })
                if medium and medium != '-':
                    levels.append({
                        'level': 4,
                        'condition': f'{medium} for {metric}',
                        'severity': 'Low',
                        'timeline': 'months'
                    })
            # Stale permissions are medium risk
            elif 'Stale' in metric:
                if high and high != '-':
                    levels.append({
                        'level': 2,
                        'condition': f'≥{high} for {metric}',
                        'severity': 'High',
                        'timeline': 'weeks'
                    })
                if medium and medium != '-':
                    levels.append({
                        'level': 3,
                        'condition': f'{medium} for {metric}',
                        'severity': 'Medium',
                        'timeline': 'months'
                    })
            # High-risk permissions are critical/high
            elif 'High Risk' in metric or 'High-Risk' in metric:
                if high and high != '-':
                    levels.append({
                        'level': 1,
                        'condition': f'≥{high} for {metric}',
                        'severity': 'Critical',
                        'timeline': 'days'
                    })
                if medium and medium != '-':
                    levels.append({
                        'level': 2,
                        'condition': f'{medium} for {metric}',
                        'severity': 'High',
                        'timeline': 'weeks'
                    })
            # Unlabeled files are compliance/baseline
            elif 'Unlabeled' in metric:
                if high and high != '-':
                    levels.append({
                        'level': 3,
                        'condition': f'≥{high} for {metric}',
                        'severity': 'Medium',
                        'timeline': 'weeks'
                    })
                if medium and medium != '-':
                    levels.append({
                        'level': 4,
                        'condition': f'{medium} for {metric}',
                        'severity': 'Low',
                        'timeline': 'months'
                    })
            # External/Anonymous sharing (not sensitive-specific) is medium
            elif 'External' in metric or 'Anonymous' in metric:
                if high and high != '-':
                    levels.append({
                        'level': 2,
                        'condition': f'≥{high} for {metric}',
                        'severity': 'High',
                        'timeline': 'weeks'
                    })
                if medium and medium != '-':
                    levels.append({
                        'level': 3,
                        'condition': f'{medium} for {metric}',
                        'severity': 'Medium',
                        'timeline': 'months'
                    })

    # Ensure we have at least one level
    if not levels:
        levels.append({
            'level': 3,
            'condition': f'{metric} threshold exceeded',
            'severity': 'Medium',
            'timeline': 'weeks'
        })

    return levels

def generate_yaml_from_excel(excel_path):
    """Generate all YAML files from Excel spreadsheet."""

    # Read the Excel file
    wb = load_workbook(excel_path)
    ws = wb.active

    category_counters = {'Data': 1, 'Identity': 1, 'Infrastructure': 1}
    category = 'Data'  # Initialize with first category

    for row in list(ws.iter_rows(values_only=True))[1:]:  # Skip header
        if not row[2]:  # Skip empty rows (check Metric column)
            continue

        category = row[1] if row[1] else category  # Category in column B (index 1)
        metric = row[2]  # Metric in column C
        measure_type = row[3]  # Measure In in column D
        low = row[4]  # Low threshold in column E
        medium = row[5]  # Medium threshold in column F
        high = row[6]  # High threshold in column G

        # Generate check ID with per-category counter
        category_prefix = {
            'Data': '1S-DATA',
            'Identity': '1S-IDENTITY',
            'Infrastructure': '1S-INFRA'
        }.get(category, '1S-UNKNOWN')

        check_id = f'{category_prefix}-{category_counters[category]:03d}'
        category_counters[category] += 1

        # Map to DRIVE platforms
        if category == 'Data':
            platform = 'SharePoint'
            pillars = ['D', 'R']
        elif category == 'Identity':
            platform = 'Active Directory'
            pillars = ['I', 'R']
        else:  # Infrastructure
            platform = 'Active Directory'
            pillars = ['R', 'V']

        # Map to levels
        levels = map_risk_to_levels(category, metric, measure_type, low, medium, high)

        # Generate YAML check
        check = {
            'check_id': check_id,
            'title': metric,
            'short_description': metric,
            'detailed_description': f'{metric} - Measured as {measure_type} with thresholds: Low={low}, Medium={medium}, High={high}',
            'category': 'Access Control' if category == 'Data' else 'Identity' if category == 'Identity' else 'Configuration',
            'platform': platform,
            'drive_pillars': pillars,
            'automatable': True,
            'owner': 'Netwrix-1Secure',
            'status': 'active',
            'created_date': '2025-10-14',
            'last_updated': datetime.now().strftime('%Y-%m-%dT%H:%M:%SZ'),
            'detection': {
                'data_sources': [
                    'Netwrix 1Secure',
                    f'{platform} State',
                    'Microsoft 365 API'
                ],
                'query_logic': f'{measure_type} measurement with thresholds',
                'data_points_required': [
                    f'1Secure metric: {metric}',
                    'Current value comparison',
                    'Threshold evaluation'
                ]
            },
            'level_thresholds': [],
            'framework_mappings': {
                'nist_csf': {'function': 'PROTECT', 'categories': [], 'controls': []},
                'cis_v8': {'controls': []},
                'cis_m365': {'version': 'v5.0', 'recommendations': []},
                'iso_27001': {'version': '2022', 'controls': []}
            },
            'remediation': {
                'automated_fix_available': False,
                '1secure_remediable': True,
                'fix_complexity': 'Medium',
                'estimated_time_minutes': 60,
                'prerequisites': ['Administrator access', '1Secure remediation access'],
                'steps': []
            },
            'validation': {
                'test_queries': [f'Verify {metric} is within acceptable threshold'],
                'false_positive_scenarios': []
            },
            'references': {
                'microsoft_docs': [],
                'industry_guidance': ['Netwrix 1Secure Risk Catalog'],
                'threat_intelligence': []
            },
            'metadata': {
                'version': '1.0.0',
                'schema_version': '2.0',
                'last_reviewed': '2025-10-14',
                'next_review_due': '2026-01-14',
                'reviewed_by': 'DRIVE-Team',
                'change_history': [{
                    'date': '2025-10-14',
                    'version': '1.0.0',
                    'change': 'Initial creation from 1Secure risk catalog',
                    'author': 'DRIVE-1Secure-Integration'
                }]
            }
        }

        # Add level thresholds (no points - binary model)
        for level_def in levels:
            threshold = {
                'level': level_def['level'],
                'threshold_id': f'{check_id}-L{level_def["level"]}',
                'threshold_condition': level_def['condition'],
                'threshold_description': f'{metric} at Level {level_def["level"]}',
                'severity': level_def['severity'],
                'business_impact': f'Impact varies based on {measure_type} threshold breach',
                'threat_timeline': f'Exploitable within {level_def["timeline"]}',
                'attacker_profile': 'Varies by severity',
                'cvss_score': 7.5 if level_def['severity'] == 'Critical' else 5.5 if level_def['severity'] == 'High' else 3.5,
                'remediation_priority': level_def['level']
            }
            check['level_thresholds'].append(threshold)

            # Add remediation step for this level
            check['remediation']['steps'].append({
                'step': level_def['level'],
                'level_target': [level_def['level']],
                'action': f'Remediate to pass Level {level_def["level"]}',
                'details': f'Use 1Secure to identify and fix: {level_def["condition"]}',
                'requires_manual': False,
                'manual_reason': 'Automated via 1Secure'
            })

        # Write YAML file
        filename = f'checks/{check_id}.yaml'
        with open(filename, 'w') as f:
            yaml.dump(check, f, default_flow_style=False, sort_keys=False, allow_unicode=True)

        print(f'✅ Created: {filename} - {metric} (Levels: {[l["level"] for l in levels]})')

    total_checks = sum(category_counters.values()) - 3  # Subtract initial values
    print(f'\n✅ Generated {total_checks} YAML check files from Excel source')
    print(f'   Data: {category_counters["Data"]-1}, Identity: {category_counters["Identity"]-1}, Infrastructure: {category_counters["Infrastructure"]-1}')

if __name__ == '__main__':
    excel_path = '/Users/jeff.warren/Downloads/1SecureRisks.xlsx'
    generate_yaml_from_excel(excel_path)
